
#__________________________import all libraries_______________________________________________________________
from selenium import webdriver
import requests
import wget
import zipfile
import os
import shutil
from tkinter import messagebox
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.microsoft import IEDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
#from selenium.webdriver import ActionChains
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.select import Select
import pandas as pd
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import argparse
from io import BytesIO
import base64
from PIL import Image
import time
import openpyxl
import pytesseract
import numpy as np
from tkinter.messagebox import *
import concurrent.futures
pytesseract.pytesseract.tesseract_cmd=r"Tesseract-OCR\tesseract.exe"

#________________________________________read excel data  ______________________________________________________

def resolve_captcha(imgs):
    text=pytesseract.image_to_string(imgs)
    txt=text.replace(" ","").replace("\n","").strip()
    return txt
    
def read_excel(file):
    commands=[]
    fields=[]
    values=[]
    
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    sheet_obj.max_row
    for i in range(2,sheet_obj.max_row+1):
        cell=sheet_obj.cell(row = i, column = 4)
        values.append(cell.value) 
    for i in range(2,sheet_obj.max_row+1):
        cell=sheet_obj.cell(row = i, column = 2)
        commands.append(cell.value)
    for i in range(2,sheet_obj.max_row+1):
        cell=sheet_obj.cell(row = i, column = 3)
        fields.append(cell.value) 
    return commands,fields,values
#___________________________________________function for perform action_________________________________________
commands,fields,values=read_excel("i3ms.xlsx")
next_commands=commands[14:]
next_fields=fields[14:]
next_values=values[14:]
permitno_list=[]
tagmorevehicleId_list=[]
vehicledetailsId_list=[]
captchaimgId=[]
entercaptchaId=[]
submitbtn=[]
tagstatusId=[]
status_data_list=[]

def drivers(browserName):
    if browserName=="Microsoft Edge":
        try:
            prog=False
            driver = webdriver.Edge(executable_path="msedgedriver.exe")   
        except:
            prog=True
            driver=None
    elif browserName=="Google Chrome":
        try:
            prog=False
            driver=webdriver.Chrome(executable_path="chromedriver.exe")
        except:
            prog=True
            driver=None
    elif browserName=="Internet Explorer":
        try:
            prog=False
            driver=webdriver.Ie(executable_path="IEDriverServer.exe")
        except:
            prog=True
            driver=None
    elif browserName=="Mozilla Firefox":
        try:
            prog=False
            driver = webdriver.Firefox(executable_path="geckodriver.exe")
        except Exception as e:
            prog=True
            driver=None
    return prog,driver

def browser(username,password,vehicleList,permitNo,browserName,permitY,permitM):
    json=[]
    try:
        for command,field,value in zip(commands,fields,values):
            if command.lower()=='gotourl' or command.lower()=="url":
                chrome_options = Options()
                #chrome_options.headless = True
                chrome_options.add_experimental_option("detach", True)
                chrome_options.add_argument("--dns-prefetch-disable")
                prog,driver=drivers(browserName)
                if prog==True:
                    try:
                        if browserName=="Microsoft Edge":
                            edge=EdgeChromiumDriverManager()
                            path=edge.install()
                            shutil.copy(path,os.getcwd())
                            driver = webdriver.Edge(executable_path=path)
                        elif browserName=="Google Chrome":
                            chrome=ChromeDriverManager()
                            path=chrome.install()
                            shutil.copy(path,os.getcwd())
                            driver = webdriver.Chrome(executable_path=path,chrome_options=chrome_options)
                        elif browserName=="Internet Explorer":
                            I=IEDriverManager()
                            path=I.install()
                            shutil.copy(path,os.getcwd())
                            driver = webdriver.Ie(executable_path=path)
                        elif browserName=="Mozilla Firefox":
                            firefox=GeckoDriverManager()
                            path=firefox.install()
                            shutil.copy(path,os.getcwd())
                            driver = webdriver.Firefox(executable_path=path)
                            
                    except Exception as e:
                        showerror("Error",f"{browserName}: something went wrong!")

                driver.get(value)
                parent_window=driver.current_window_handle
               
                # except Exception as e:
                #     print("exception as e",e)
                    # url = 'https://chromedriver.storage.googleapis.com/LATEST_RELEASE'
                    # response = requests.get(url)
                    # version_number = response.text
                    # download_url = "https://chromedriver.storage.googleapis.com/" + version_number +"/chromedriver_win32.zip"
                    # latest_driver_zip = wget.download(download_url,os.getcwd())
                    # with zipfile.ZipFile(latest_driver_zip, 'r') as zip_ref:
                    #     zip_ref.extractall()
                    # os.remove(latest_driver_zip)
                    # driver=webdriver.Chrome(chrome_options=chrome_options)
                    # driver.get(value) 
                
            elif command.lower()=="username":
                driver.find_element(By.XPATH,field).send_keys(username)
                
            elif command.lower()=="password":
                driver.find_element(By.XPATH,field).send_keys(password)
 
            elif command.lower()=='entertext' or command.lower()=="textenter" or command.lower()=="input":
                try:
                    driver.find_element(By.XPATH,field).send_keys(value)
                except:
                    pass
                
            elif command.lower()=="selectyear":
                try:
                    sel = Select(driver.find_element_by_xpath(field))
                    sel.select_by_visible_text(permitY)
                except:
                    pass
                
                       
            elif command.lower()=="selectmonth":
                try:
                    sel = Select(driver.find_element_by_xpath(field))
                    sel.select_by_visible_text(permitM)
                except:
                    pass
                
            elif command.lower()=="selecttransporter":
                try:
                    # driver.find_element(By.XPATH,field).click()
                    btn=driver.find_element(By.XPATH,field)
                    ActionChains(driver).move_to_element(btn).click(btn).perform() 
                except Exception as e:
                    pass

            elif  command.lower()=='clickbutton' or command.lower()=="click" or command.lower()=="buttonclick" or command.lower()=="checkbutton" or command.lower()=="checkbox" or command.lower()=="radiobutton" or command.lower()=="radiobox" or command.lower()=="button":
                try:
                    driver.find_element(By.XPATH,field).click()
                except:
                    pass

            elif command.lower()=='switchtoframe':
                try:
                    driver.switch_to.frame(driver.find_element(By.XPATH,field))
                except Exception as e:
                    print("switch to frame ",e)

            elif command.lower()=="parentframe":
                try:
                    driver.switch_to.default_content()
                except Exception as e:
                    print("switch to default",e)
                    pass

            elif command.lower()=="newtab":
                try:
                    if value:
                        driver.execute_script(f"window.open('about:blank', '{value}');")
                        driver.switch_to.window(value)
                        driver.get(value)
                    else:
                        child_window=driver.window_handles
                        for child in child_window:
                            if parent_window!=child:
                                driver.switch_to.window(child)
                except:
                    pass
                            
            elif command.lower()=="parentwindow":
                try:
                    driver.switch_to.window(parent_window)
                except Exception as e:
                    print("parent window: ",e)
                    pass

            elif command.lower()=='getpermitno':
                try:
                    i=0
                    permit_list=driver.find_elements(By.XPATH,field)
                    a=3
                    for permit in permit_list:
                        if i==a:
                            permitno_list.append(permit.text)
                            a=a+8
                        i+=1
                except:
                    pass
                                          
            elif command.lower()=="gettagmorevehicleid":
                try:
                    action_list=driver.find_elements(By.XPATH,field)
                    for action in action_list:
                        tagmorevehicleId_list.append(action.get_attribute('id'))
                except:
                    pass
                               
            elif command.lower()=="tagmorevehicleclick":
                try:
                    permitaction_list=list(zip(permitno_list,tagmorevehicleId_list))
                    clickid=[]
                    for permitaction in permitaction_list:
                        if permitNo in permitaction:
                            clickid.append(permitaction[1])
                            break
                    if len(clickid)!=0:
                        driver.find_element(By.XPATH,f"//a[@id='{clickid[0]}']").click()      
                except Exception as e:
                    driver.quit()
                     
            elif command.lower()=="solvecaptcha":
                try:
                    captchaimgId.append(field)
                except:
                    pass
            elif command.lower()=="entercaptcha":
                try:
                    entercaptchaId.append(field)
                except:
                    pass
                
            elif command.lower()=="submitbtn":
                try:
                    submitbtn.append(field)
                except:
                    pass
           
            elif command.lower()=="vehicleno":
                for i in vehicleList:
                    try:
                        for next_command,next_field,next_value in zip(next_commands,next_fields,next_values):
                            if next_command.lower()=="vehicleno":
                                try:
                                    driver.find_element(By.XPATH,next_field).clear()
                                    driver.find_element(By.XPATH,next_field).send_keys(i)
                                except:
                                    driver.find_element(By.XPATH,next_field).clear()
                                    driver.find_element(By.XPATH,next_field).send_keys(i)
 
                            elif  next_command.lower()=='clickbutton' or next_command.lower()=="click" or next_command.lower()=="buttonclick" or next_command.lower()=="checkbutton" or next_command.lower()=="checkbox" or next_command.lower()=="radiobutton" or next_command.lower()=="radiobox" or next_command.lower()=="button":
                                try:
                                    btn=driver.find_element(By.XPATH,next_field)
                                    ActionChains(driver).move_to_element(btn).click(btn).perform() 
                                except Exception as e:
                                    pass
                            elif next_command.lower()=="status":
                                status=driver.find_element(By.XPATH,next_field)
                                if status.text:
                                    dict={"vehicleno":i,
                                        "permitno":permitNo,
                                        "status":[status.text]}
                                    json.append(dict)
                                    driver.back()
                                    break
                            
                            elif next_command.lower()=="solvecaptcha":
                                captchaimgId.clear()
                                captchaimgId.append(next_field)
                                
                            elif next_command.lower()=="entercaptcha":
                                entercaptchaId.clear()
                                entercaptchaId.append(next_field)
                                
                            elif next_command.lower()=="submitbtn":
                                submitbtn.clear()
                                submitbtn.append(next_field)

                            elif next_command.lower()=="checkcaptcha":
                                captcha_lists=list(zip(captchaimgId,entercaptchaId,submitbtn))
                                while True:
                                    try:
                                        for captchaimgid,entercaptchaid,submitbutton in captcha_lists:
                                            element = driver.find_element_by_xpath(captchaimgid).screenshot_as_base64
                                            elem = driver.find_element_by_xpath(captchaimgid)
                                            action_chain = ActionChains(driver)
                                            action_chain.move_to_element(elem)
                                            action_chain.perform()
                                            
                                            img = Image.open(BytesIO(base64.b64decode(element)))
                                            img_array=np.array(img)
                                            
                                            txt=resolve_captcha(img_array)
                                            captcha_list=txt.split("?")
                                        
                                            if captcha_list[0].endswith("firstNo."):
                                                firstNo=captcha_list[1].replace("=","").split(",")[0]
                                                firstNo=int(firstNo)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(firstNo)
                                                break  
                                            elif captcha_list[0].endswith("lastNo."):
                                                lastNo=captcha_list[1].replace("=","").split(",")[-1]
                                                lastNo=int(lastNo)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(lastNo)
                                                break
                                            elif captcha_list[0].endswith("MiddleNo."):
                                                middleNo=captcha_list[1].replace("=","").split(",")[1]
                                                middleNo=int(middleNo)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(middleNo)
                                                break
                                            elif captcha_list[0].endswith("greatestNo."):
                                                value_list=captcha_list[1].replace("=","").split(",")
                                                No=[int(i) for i in value_list]
                                                greatestNo=max(No)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(greatestNo)
                                                break
                                            elif captcha_list[0].endswith("smallestNo."):
                                                value_list=captcha_list[1].replace("=","").split(",")
                                                No=[int(i) for i in value_list]
                                                smallestNo=min(No)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(smallestNo)
                                                break
                                            else:
                                                expression=captcha_list[1].replace("=","")
                                                solution=eval(expression)
                                                driver.find_element(By.XPATH,entercaptchaid).clear()
                                                driver.find_element(By.XPATH,entercaptchaid).send_keys(solution)
                                                break
                                        try:
                                            btn2=driver.find_element(By.XPATH,submitbutton)
                                            ActionChains(driver).move_to_element(btn2).click(btn2).perform()
                                            alert=driver.switch_to_alert()
                                            alert_text=alert.text
                                            if alert_text:
                                                dict={"vehicleno":i,
                                                    "permitno":permitNo,
                                                    "status":["VEHICLE TAGGED SUCCESSFULLY"]}
                                                json.append(dict)
                                                WebDriverWait(driver, 10).until(EC.alert_is_present())
                                                driver.switch_to.alert.accept()
                                                break  
                                            else:
                                                driver.refresh()
                                                            
                                        except Exception as e:
                                            driver.refresh()
                                            
                                    except Exception as e:
                                        driver.refresh()
                                        
                    except Exception as e:
                        dict={"vehicleno":i,
                              "permitno":permitNo,
                              "status":["VEHICLE NOT TAGGED"]}
                        json.append(dict)
                        btn3=driver.find_element(By.XPATH,"//input[@id='btnCancel']")
                        ActionChains(driver).move_to_element(btn3).click(btn3).perform()
                        driver.refresh()
        driver.quit()                                                              
    except Exception as e:
        driver.quit()
    return json

# json=browser("AAFCG2536F4","Gargsons@1234",[1,1,1,1],"I122100995","Mozilla Firefox")
# print(json)

