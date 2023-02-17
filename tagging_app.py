from tkinter import*
import os
import time
from PIL import ImageTk,Image
from tkinter import ttk
from click import option
import openpyxl as xl
from i3ms import browser,drivers
from tkinter import messagebox
import pandas as pd
from threading import Thread
import arrow as dt
from tkinter.filedialog import askopenfilename,asksaveasfilename
datetime=dt.now()
import requests
#__________________________________________________api calls_______________________________________________________
#response=requests.get("https://lms.tranzol.com/Tranzol/ValidateLicence",params={"C":"santuka","P":104})

#________________________________________________________main app_________________________________________________

slno=1
def vehicle_tagging():
    vehicle_list=[]
    maindata=[]
    tagging_win=Tk()
    tagging_win.title("Vehicle Tagging")
    tagging_win.iconbitmap("img/truck.ico")
    width=730
    height=480
    start_window_from_x_axis=(tagging_win.winfo_screenwidth()//2)-(width//2)
    start_window_from_y_axis=(tagging_win.winfo_screenheight()//2)-(height//2)
    tagging_win.geometry(f"{width}x{height}+{start_window_from_x_axis}+{start_window_from_y_axis}")
    tagging_win.resizable(0,0)

    def clear():
        #user_var.set("")
        password_var.set("")
        permit_var.set("")
        vehicle_var.set("")
        browser_name.set("Google Chrome")
        permitYr_var.set(datetime.format("YYYY"))
        permitMnth_var.set(datetime.format("MMM"))
        password_entry.focus_force()

    def open_file():
        vehicle_list.clear()
        filetypes = (
            ('Excel file', '*.xlsx'),
        )

        filename = askopenfilename(
            title='Open a file',
            initialdir='/',
            filetypes=filetypes)

        if filename:
            file=filename.split("/")
            vehicle_var.set(file[-1])
            wb_obj = xl.load_workbook(filename)
            sheet_obj = wb_obj.active
            sheet_obj.max_row
            for i in range(1,sheet_obj.max_row+1):
                cell=sheet_obj.cell(row = i, column = 2)
                vehicle_list.append(cell.value)

    def progressbar(browser_n):
        l=Label(text="webdriver installing...",font=("arial 10"))
        s = ttk.Style()
        progress=ttk.Progressbar(style="red.Horizontal.TProgressbar", length=170,mode="indeterminate")
        s.configure("red.Horizontal.TProgressbar", foreground='red', background='green')
        prog,_=drivers(browser_n)
        try:
            _.quit()
        except:
            pass
        if prog==True:
            l.place(x=24,y=455)
            progress.place(x=167,y=457)
            progress.start(10) 
            time.sleep(10)
            progress.stop()
            progress.place_forget()
            l.place_forget()

    def tag_vehicle1(username,password,vehicle_list,permitNo,browser_n,permitY,permitM):
        global slno
        clear()
        data=browser(username,password,vehicle_list,permitNo,browser_n,permitY,permitM)     
        maindata.extend(data)
        for d in data:
            table.insert('',"end",text=slno,values=(d['permitno'],d['vehicleno'],d['status'][0]))
            slno=slno+1       
            
    def tag_vehicle(e):
        username=user_var.get()
        password=password_var.get()
        permitNo=permit_var.get()
        if username and password and permitNo and vehicle_list:
            question=messagebox.askyesno("Tag Vehicle","Do you want tag vehicle?")
            if question:
                x = Thread(target=tag_vehicle1, args=(username,password,vehicle_list,permitNo,browser_name.get(),permitYr_var.get(),permitMnth_var.get()))
                y = Thread(target=progressbar,args=(browser_name.get(),))
                x.start()
                y.start()
        else:
            messagebox.showerror("Error","Please fill all fields")

    def excel():
        dt=pd.DataFrame(data=[[f"{i['permitno']}", f"{i['vehicleno']}",f"{i['status'][0]}"] for i in maindata],columns=["Permit No.","Vehicle No.","Status"])
        if messagebox.askyesno("File save","Do you want to save file?"):
            filetypes=(
                ("Excel file", "*.xlsx"),
                ) 
            file_name = asksaveasfilename(filetypes=filetypes, defaultextension='.xlsx')
            if file_name:
                dt.to_excel(file_name)

#____________________________UI part_______________________________________________________________
    
    top_frame=Frame(tagging_win,bg="#0095FF")
    top_frame.place(x=0,y=0,width=730,height=90)

    load_top_img=Image.open("img/truck1.png")
    resize_top_img=load_top_img.resize((100,60))
    top_img=ImageTk.PhotoImage(resize_top_img)
    top_img_label=Label(top_frame,image=top_img,bg="#0095FF",fg="#0095FF")
    top_img_label.place(x=15,y=5)

    title_label=Label(top_frame,text="VEHICLE TAGGING SYSTEM",font=("arial 14"),bg="#0095FF",fg="white")
    title_label.pack(pady=10)

    screen_label=Label(top_frame,text="Tag Vehicle",font=("arial 12 bold"),bg="#0095FF",fg="white")
    screen_label.pack(pady=4)

    left_frame=Frame(tagging_win,bg="white")
    left_frame.place(x=0,y=90,width=400,height=360)

    username_label=Label(left_frame,text="UserID",font=("arial 13"),bg="white")
    username_label.place(x=25,y=45)

    user_var=StringVar()
    user_entry=Entry(left_frame,font=("arial 12"),textvariable=user_var,bd=2,state='disabled')
    user_entry.place(x=150,y=45,width=200)
    user_var.set("AAEFS6131L14")
    
    password_label=Label(left_frame,text="Password",font=("arial 13"),bg="white")
    password_label.place(x=25,y=83)

    password_var=StringVar()
    password_entry=Entry(left_frame,font=("arial 12"),show="\u2022",textvariable=password_var,bd=2)
    password_entry.place(x=150,y=83,width=200)
    
    permit_label=Label(left_frame,text="Permit No",font=("arial 13"),bg="white")
    permit_label.place(x=25,y=120)

    permit_var=StringVar()
    permit_entry=Entry(left_frame,font=("arial 12"),textvariable=permit_var,bd=2)
    permit_entry.place(x=150,y=120,width=200)

    permit_yr=Label(left_frame,text="Permit Year",font=("arial 12"),bg="white")
    permit_yr.place(x=25,y=155)

    permitYr_var=StringVar()
    years=(datetime.format("YYYY"),datetime.shift(years=-1).format("YYYY"))
    permitYr_box=OptionMenu(left_frame,permitYr_var, *years)
    permitYr_box.place(x=150,y=155,height=25)
    permitYr_var.set(datetime.format("YYYY"))
    permitYr_box["menu"].config(font=("arial 12"))
    permitYr_box.config(font=("arial 11"))

    permit_mnth=Label(left_frame,text="Month",font=("arial 12"),bg="white")
    permit_mnth.place(x=230,y=155)

    permitMnth_var=StringVar()
    months=(datetime.format("MMM"),datetime.shift(months=-1).format("MMM"))
    permitMnth_box=OptionMenu(left_frame,permitMnth_var, *months)
    permitMnth_box.place(x=279,y=155,height=25)
    permitMnth_box.config(font=("arial 11"))
    permitMnth_box["menu"].config(font=("arial 12"))
    permitMnth_var.set(datetime.format("MMM"))

    vehicle_label=Label(left_frame,text="Vehicle No",font=("arial 13"),bg="white")
    vehicle_label.place(x=25,y=192)

    vehicle_var=StringVar()
    vehicle_entry=Entry(left_frame,font=("arial 12"),textvariable=vehicle_var,bd=2,width=17,state="disable")
    vehicle_entry.place(x=150,y=194,width=200)

    load_img=Image.open("img/open-folder.png")
    resize_img=load_img.resize((30,30))
    fileimg=ImageTk.PhotoImage(resize_img)

    file_btn=Button(left_frame,image=fileimg,relief=FLAT,bd=2,bg="white",command=open_file)
    file_btn.place(x=318,y=187)

    browser_label=Label(left_frame,text="Browser",font=("arial 13"),bg="white")
    browser_label.place(x=25,y=231)

    browser_name=StringVar()
    options=["Google Chrome","Microsoft Edge"]
  
    browser_name.set("Google Chrome")
    browser_btn=OptionMenu(left_frame,browser_name, *options)
    browser_btn.config(width=14,font=("arial 11"))
    browser_btn["menu"].config(font=("arial 12"))
    browser_btn.place(x=150,y=230)
    
    clear_btn=Button(left_frame,text="Clear",font=("arial",12,"bold"),bd=2,relief=RIDGE,width=13,bg="red",fg="white",command=clear)
    clear_btn.place(x=25,y=290)

    tag_btn=Button(left_frame,text="Tag Vehicle",font=("arial",12,"bold"),bd=2,relief=RIDGE,width=13,bg="#0095FF",fg="white",command=lambda:tag_vehicle("event"))
    tag_btn.place(x=210,y=290)

#______________________________right frame________________________________________________
    right_frame=Frame(tagging_win,bg="white",relief=RIDGE,bd=3)
    right_frame.place(x=370,y=90,width=359,height=391)

    xls_img=Image.open("img/excel.png")
    xls_img=xls_img.resize((25,25))
    xls_img=ImageTk.PhotoImage(xls_img)

    report_label=Label(right_frame,text=f"Report",font=("arial 12"),bg="white")
    report_label.place(x=10,y=10)

    print_btn=Button(right_frame,image=xls_img,relief=FLAT,bg="white",command=excel)
    print_btn.place(x=310,y=4)
    
 #______________________________________Gui Table ____________________________________________
    
    style=ttk.Style(tagging_win)
    style.theme_use("alt")
    style.configure('mystyle.Treeview.Heading',font=('arial',10,"bold"),foreground="white",background="gray",activebackground="gray",relief=RIDGE,borderwidth=1)  # for heading
    style.configure('mystyle.Treeview',font=('arial',10),relief=RIDGE,borderwidth=3)  # for data
    table=ttk.Treeview(right_frame,style='mystyle.Treeview')
    table.place(x=0,y=42,width=350,height=330)
    table.tag_configure("tag")
        
    table['columns']=('1','2','3')
    table.column('#0',width=50,minwidth=25,anchor="center")
    table.column('1',width=120,minwidth=50,anchor="center")
    table.column('2',width=120,minwidth=25,anchor="center")
    table.column('3',width=350,minwidth=25,anchor="center")
    #column heading
    table.heading('#0',text='Sl No')
    table.heading('1',text='Permit No')
    table.heading('2',text='Vehicle No')
    table.heading('3',text='Status')
    
    scroll_table_horizontal=Scrollbar(right_frame,orient=HORIZONTAL)
    scroll_table_horizontal.pack(side=BOTTOM,fill=X)
    scroll_table_vertical=Scrollbar(right_frame,orient=VERTICAL)
    scroll_table_vertical.place(x=340,y=63,height=306)
    
    scroll_table_vertical.config(command=table.yview)
    scroll_table_horizontal.config(command=table.xview)
    
    table.config(xscrollcommand=scroll_table_horizontal.set)
    table.config(yscrollcommand=scroll_table_vertical.set)

    #user_entry.bind("<Return>",lambda event:password_entry.focus_force())
    password_entry.bind("<Return>",lambda event:permit_entry.focus_force())
    permit_entry.bind("<Return>",lambda event:vehicle_entry.focus_force())
    vehicle_entry.bind("<Return>",tag_vehicle)
    password_entry.focus_force()
    tagging_win.mainloop()

vehicle_tagging()
